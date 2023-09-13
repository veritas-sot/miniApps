import logging
import os
import json
import csv
import hashlib
import yaml
import pytricia
from pathlib import Path
from openpyxl import load_workbook
from slugify import slugify


def calculate_checksum(row):
    data = ""
    for key, value in row.items():
        if isinstance(value, list):
            my_list = ''.join(value)
            data += my_list
        elif value is None:
            pass
        else:
            data += value
    # calculate md5 sum
    return hashlib.md5(data.encode('utf-8')).hexdigest()

def build_dict(value, my_dict, keys):
    if isinstance(keys, list):
        my_dict[keys[0]] = build_dict(value, my_dict, keys[1])
    else:
        return {keys: value}
    return my_dict

def get_device_defaults(prefixe, ip):
    if prefixe is None:
        return {}

    prefix_path = get_prefix_path(prefixe, ip)
    defaults = {}
    for prefix in prefix_path:
        defaults.update(prefixe[prefix])

    return defaults

def get_prefix_path(prefixe, ip):
    prefix_path = []
    pyt = pytricia.PyTricia()

    # build pytricia tree
    for prefix_ip in prefixe:
        pyt.insert(prefix_ip, prefix_ip)

    prefix = pyt.get(ip)
    prefix_path.append(prefix)

    parent = pyt.parent(prefix)
    while (parent):
        prefix_path.append(parent)
        parent = pyt.parent(parent)
    return prefix_path[::-1]

def parse_row(row, value_mapping, kobold_config):
    data = {}
    # the next mapping is used to map column names to SOT names
    mappings = kobold_config['mappings'].get('import')
    # prepare data
    for k, value in row.items():
        if k in mappings:
            key = mappings.get(k)
        else:
            key = k
        # value_mappings are uses to 'rename' values
        if value_mapping:
            mapping = value_mapping.get(key)
            if mapping:
                for map_key,map_value in mapping.items():
                    if value == map_key:
                        value = map_value

        if key.startswith('cf_'):
            if 'custom_fields' not in data:
                data['custom_fields'] = {}
            data['custom_fields'][key.split('cf_')[1]] = value
        elif '__' in key:
            # check if key has subkeys
            path = key.split('__')
            d = build_dict(value, {}, path)
            if path[0] == 'interfaces':
                data.update(d[path[0]])
            else:
                if path[0] not in data:
                    data[path[0]] = {}
                data[path[0]].update(d[path[0]])
        else:
            data[key] = value
    return data

def read_csv(args, kobold_config, value_mapping=None):
    contains_interface = False
    data = []

    filename = args.data
    delimiter = kobold_config['defaults']['import'].get('delimiter',',')
    quotechar = kobold_config['defaults']['import'].get('quotechar','|')
    quoting_cf = kobold_config['defaults']['import'].get('quoting','minimal')
    newline = kobold_config['defaults']['import'].get('newline','')
    if quoting_cf == "none":
        quoting = csv.QUOTE_NONE
    elif quoting_cf == "all":
        quoting = csv.QUOTE_ALL
    elif quoting_cf == "nonnumeric":
        quoting = csv.QUOTE_NONNUMERIC
    else:
        quoting = csv.QUOTE_MINIMAL
    logging.info(f'reading {filename} delimiter={delimiter} quotechar={quotechar} newline={newline} quoting={quoting_cf}')

    # read CSV file
    with open(filename, newline=newline) as csvfile:
        csvreader = csv.DictReader(csvfile, 
                                   delimiter=delimiter, 
                                   quoting=quoting,
                                   quotechar=quotechar)
        for row in csvreader:
            # check if we have interfaces to update or import
            for name in row:
                if 'interface' in name:
                    contains_interface = True

            old_checksum = new_checksum = 0
            old_checksum = new_checksum = 0
            if 'checksum' in row:
                old_checksum = row['checksum']
                del row['checksum']
                new_checksum = calculate_checksum(row)

            if args.update and old_checksum != new_checksum:
                data.append(parse_row(row, value_mapping, kobold_config))
            elif args.force or args.import_data:
                data.append(parse_row(row, value_mapping, kobold_config))

    return contains_interface, data

def read_xlsx(args, kobold_config, value_mapping=None):
    contains_interface = False
    data = []
    table = []

    # Load the workbook
    workbook = load_workbook(filename = args.data)
    # Select the active worksheet
    worksheet = workbook.active
    
    # loop through table and build list of dict
    rows = worksheet.max_row
    columns = worksheet.max_column
    for row in range(2, rows + 1):
        line = {}
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)

    for row in table:
        if 'interface' in row:
            contains_interface = True
        old_checksum = new_checksum = 0
        old_checksum = new_checksum = 0
        if 'checksum' in row:
            old_checksum = row['checksum']
            del row['checksum']
            new_checksum = calculate_checksum(row)

        if args.update and old_checksum != new_checksum:
            data.append(parse_row(row, value_mapping, kobold_config))
        elif args.force or args.import_data:
            data.append(parse_row(row, value_mapping, kobold_config))
    return contains_interface, data

def read_json(args, kobold_config, value_mapping=None):
    data = []
    filename = args.data

    logging.debug(f'reading HLDM in {args.data}')
    for path in Path(args.data).rglob('*.json'):
        with open(path, 'r') as f:
            data.append(json.load(f))

    return True, data

def bulk_update(sot, data, kobold_config, endpoint):
    nb = sot.rest(url=kobold_config['sot']['nautobot'], token=kobold_config['sot']['token'])
    nb.session()
    response = nb.patch(url=f"api/{endpoint}/", json=data)
    if response.status_code != 200:
        logging.error(f'could not update data; got error {response.content}')
    else:
        logging.info(f'data updated')

def prepare_new_data(args, sot, defaults, new_data):
    device_defaults = {}
    data = []

    for item in new_data:
        if 'name' in item:
            hostname = item.get('name').lower()
            del item['name']
        else:
            raise Exception('need hostname to add data')

        field = None
        if 'primary_ip' in item:
            field = 'primary_ip'
        elif 'primary_ip4' in item:
            field = 'primary_ip4'
        else:
            raise Exception('need IP address to add data')

        # the used IP address is stored in 'primary_ip4
        primary_ip = item.get(field)
        del item[field]
        item['primary_ip4'] = primary_ip

        device_defaults = get_device_defaults(defaults, primary_ip)
        device_properties = {
            "name": hostname,
            "site": {'slug': slugify(device_defaults.get('site'))},
            "device_role": {'slug': slugify(device_defaults.get('device_role'))},
            "device_type": {'slug': slugify(device_defaults.get('device_type'))},
            "manufacturer": {'slug': slugify(device_defaults.get('manufacturer'))},
            "platform": {'slug': slugify(device_defaults.get('platform'))},
            "status": device_defaults.get('status','active'),
            "custom_fields": device_defaults.get('custom_fields',{})
        }

        # customfields are special; we have to merge the dict
        if 'custom_fields' in item:
            cfields = item['custom_fields']
            del item['custom_fields']

        # overwrite existing values with import
        device_properties.update(item)
        # and add custom fields of the item
        for key, value in cfields.items():
            device_properties['custom_fields'][key] = value
        data.append(device_properties)

    return data

def convert_hldm_to_properties(args, sot, data):

    response = []
    for hldm in data:
        properties = {}
        for key, value in hldm.items():
            if key == 'custom_fields':
                properties['custom_fields'] = {}
                for k,v in value.items():
                    properties['custom_fields'][k] = v
            elif key == 'platform':
                manufacturer = value.get('manufacturer').get('slug')
                del value['manufacturer']
                properties['platform'] = {'slug': value.get('slug')}
                properties['manufacturer'] = {'slug': manufacturer}
            elif key == 'primary_ip4':
                if 'address' in value and value['address']:
                    properties['primary_ip4'] = value['address']
                else:
                    properties['primary_ip4'] = None
            elif key == 'device_role':
                properties['device_role'] = {'slug': value.get('slug')}
            elif key == 'device_type':
                properties['device_type'] = {'slug': value.get('slug')}
            elif key == 'status':
                properties['status'] = value.get('slug')
            else:
                properties[key] = value

        response.append(properties)
    return response

def add_device_from_hldm(sot, hostname, device_properties):

    interfaces = device_properties.get('interfaces')
    config_context = device_properties.get('config_context')
    primary_ip4 = device_properties.get('primary_ip4')
    del device_properties['interfaces']
    del device_properties['config_context']
    del device_properties['primary_ip4']

    nb_device = sot.device(hostname) \
                   .use_defaults(True) \
                   .return_device(True) \
                   .add(device_properties)
    # now add interfaces
    logging.debug(f'adding interfaces to {hostname}')

    lag = []
    physical = []
    virtual = []

    for interface in interfaces:
        interface['device'] = {'name': hostname}
        if interface['type'] == 'LAG':
            interface['type'] = slugify(interface['type'])
            interface['mode'] = slugify(interface['mode'])
            lag.append(interface)
        elif interface['type'] == 'VIRTUAL':
            interface['type'] = slugify(interface['type'])
            interface['mode'] = slugify(interface['mode'])
            if 'untagged_vlan' in interface and interface['untagged_vlan']:
                vlan = interface['untagged_vlan']
                vlan_id = sot.ipam.vlan(vlan['vid']).get()
                if not vlan_id:
                    vlan['status'] = 'active'
                    logging.debug('adding VLAN {vlan["vid"]}')
                    vlan_id = sot.ipam.vlan(vlan['vid']).add(vlan)
                    interface['untagged_vlan'] = vlan_id.id
            if 'tagged_vlans' in interface and interface['tagged_vlans']:
                vlans = interface['tagged_vlans']
                untagged_vlan = []
                for vlan in vlans:
                    vlan_id = sot.ipam.vlan(vlan['vid']).get()
                    if not vlan_id:
                        vlan['status'] = 'active'
                        logging.debug('adding VLAN {vlan["vid"]}')
                        vlan_id = sot.ipam.vlan(vlan['vid']).add(vlan)
                    untagged_vlan.append(vlan_id.id)
                interface['tagged_vlans'] = untagged_vlan
            virtual.append(interface)
        else:
            interface['type'] = slugify(interface['type'].replace('A_',''))
            physical.append(interface)

    sot.device(hostname).add_list_of_interfaces(lag)
    sot.device(hostname).add_list_of_interfaces(virtual)
    sot.device(hostname).add_list_of_interfaces(physical)

    for interface in interfaces:
        interface_name = interface.get('name')
        if 'ip_addresses' in interface and len(interface['ip_addresses']) > 0:
            addr = interface['ip_addresses'][0].get('address')
            if addr:
                logging.debug(f'adding {addr} to SOT')
                nb_addr = sot.ipam.ipv4(addr).add({'status': 'active'})
                logging.debug(f'assigning {interface_name} on {nb_device} to {addr}')
                assigned_interface = sot.ipam \
                                        .assign(interface_name) \
                                        .on(nb_device) \
                                        .to(addr)
                if addr == primary_ip4:
                    logging.debug(f'primary ip {addr} found on {interface_name}')
                    success = nb_device.update({'primary_ip4': nb_addr.id})

def import_data(args, sot, kobold_config):

    value_mapping = None
    # maybe we need some default values to import devices.
    # Get default values of prefixes. This is needed only once
    name_of_repo = kobold_config['git']['defaults']['repo']
    path_to_repo = kobold_config['git']['defaults']['path']
    filename = kobold_config['git']['defaults']['filename']
    default_repo = sot.repository(repo=name_of_repo, path=path_to_repo)
    defaults_str = default_repo.get(filename)
    if defaults_str is None:
        logging.error("could not load defaults")
        raise Exception('could not load defaults')
    try:
        defaults_yaml = yaml.safe_load(defaults_str)
        if defaults_yaml is not None and 'defaults' in defaults_yaml:
            defaults = defaults_yaml['defaults']
    except Exception as exc:
        logging.critical("Cannot read default values; got exception: %s" % exc)
        raise Exception("cannot read default values")

    # try to build a dict from the default values
    defaults_yaml = yaml.safe_load(defaults_str)
    if defaults_yaml is not None and 'defaults' in defaults_yaml:
        defaults = defaults_yaml['defaults']

    # read value mapping
    if args.value_mapping:
        with open(args.value_mapping,'r') as f:
            value_mapping = yaml.safe_load(f.read())['values']

    filename = args.data
    logging.debug(f'reading {filename}')
    if 'csv' in filename:
        contains_interface, data = read_csv(args, kobold_config)
    elif 'xlsx' in filename:
        contains_interface, data = read_xlsx(args, kobold_config)
    else:
        contains_interface, data = read_json(args, kobold_config)

    if 'csv' in filename or 'xlsx' in filename:
        properties = prepare_new_data(args, sot, defaults, data)
    else:
        properties = convert_hldm_to_properties(args, sot, data)

    for device_properties in properties:
        hostname = device_properties.get('hostname')
        primary_ip4 = device_properties.get('primary_ip4')
        if 'interface_name' in device_properties:
            primary_interface = device_properties.get('interface_name')
            del device_properties['interface_name']
        else:
            primary_interface = kobold_config['defaults']['import'].get('primary_interface','primary_interface')
        logging.info(f'adding {hostname} / {primary_interface} {primary_ip4}')

        # add device
        if 'csv' in filename or 'xlsx' in filename:
            # remove unnecessary properties
            del device_properties['interfaces']
            del device_properties['config_context']
            del device_properties['primary_ip4']
            device = sot.device(hostname) \
                        .use_defaults(True) \
                        .return_device(True) \
                        .primary_interface(primary_interface) \
                        .primary_ipv4(primary_ip4) \
                        .make_primary(True) \
                        .add(device_properties)
        else:
            add_device_from_hldm(sot, hostname, device_properties)

def update_data(args, sot, kobold_config):
    logging.info(f'update data')
    if 'csv' in args.data:
        contains_interface, data = read_csv(args, None, kobold_config)
    elif 'xlsx' in args.data:
        contains_interface, data = read_xlsx(args, None, kobold_config)
    if len(data) > 0:
        # are we able to make a bulk update by using the ID?
        if 'id' in data[0]:
            if contains_interface:
                bulk_update(sot, data, kobold_config, "dcim/interfaces")
            else:
                bulk_update(sot, data, kobold_config, "dcim/devices")
        else:
            # we do not have an ID; use hostname/interface name instead
            if contains_interface:
                for interface in data:
                    sot.device(interface['hostname']).interface(interface['name']).update(interface)
            else:
                for device in data:
                    sot.device(device['hostname']).update(device)