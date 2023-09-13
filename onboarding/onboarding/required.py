import re
import yaml
import logging
import os
import glob
import json
import csv
from slugify import slugify
from openpyxl import load_workbook
from veritas.sot import sot as sot


# global cache
_global_cache = {}


def required(sot, device_defaults, device_facts, ciscoconf, onboarding_config):

    basedir = "%s/%s" % (onboarding_config.get('git').get('app_configs').get('path'),
                         onboarding_config.get('git').get('app_configs').get('subdir'))
    directory = os.path.join(basedir, './onboarding/required/')
    files = []

    """
    Add the beginning response is empty. At the end this dict contains the additional values that are added to nautobot
    """
    response = {}

    logging.debug(f'reading config from {directory} files for creating required values')
    # we read all *.yaml files in our required data config dir
    for filename in glob.glob(os.path.join(directory, "*.yaml")):
        logging.debug(f'reading {filename}')

        config = read_file(filename, device_defaults)
        if config is None:
            continue

        # add filename to our list of files that were processed
        files.append(os.path.basename(filename))

        for item_config in config.get('required'):
            if 'file' in item_config:
                get_additional_values(response, 
                                      item_config,
                                      device_facts,
                                      device_defaults,
                                      onboarding_config)
            else:
                process_matches(response,
                                device_facts, 
                                device_defaults, 
                                item_config, 
                                ciscoconf)

    return response

def get_additional_values(response, item_config, device_facts, device_defaults, onboarding_config):
    file_format = item_config.get('format','csv')

    if file_format == 'csv':
        return add_values_from_csv(response, item_config, device_facts, device_defaults, onboarding_config)
    elif file_format == 'excel':
        return add_values_from_excel(response, item_config, device_facts, device_defaults, onboarding_config)
    else:
        logging.error(f'unknown file format {file_format}')
        return response

def add_values_from_csv(response, item_config, device_facts, device_defaults, onboarding_config):

    basedir = "%s/%s" % (onboarding_config.get('git').get('app_configs').get('path'),
                         onboarding_config.get('git').get('app_configs').get('subdir'))
    directory = os.path.join(basedir, './onboarding/required/')
    
    filename = "%s/%s" % (directory, item_config.get('file'))
    logging.debug(f'reading additional values from {filename}')

    # set default values
    delimiter = item_config.get('delimiter',',')
    quotechar = item_config.get('quotechar','|')
    quoting_cf = item_config.get('quoting','minimal')
    newline = item_config.get('newline','')
    if quoting_cf == "none":
        quoting = csv.QUOTE_NONE
    elif quoting_cf == "all":
        quoting = csv.QUOTE_ALL
    elif quoting_cf == "nonnumeric":
        quoting = csv.QUOTE_NONNUMERIC
    else:
        quoting = csv.QUOTE_MINIMAL
    logging.info(f'reading mapping {filename} delimiter={delimiter} quotechar={quotechar} newline={newline} quoting={quoting_cf}')

    # read CSV file
    with open(filename, newline=newline) as csvfile:
        csvreader = csv.DictReader(csvfile, delimiter=delimiter, quoting=quoting, quotechar=quotechar)
        for row in csvreader:
            # maybe there are multiple items
            for matches_on in item_config.get('matches_on'):
                # sot key => name of key in our sot
                # csv_key => name of column in our csv file
                for sot_key, csv_key in matches_on.items():
                    df = device_facts.get(sot_key)
                    if df and df == row.get(csv_key):
                        del row[csv_key]
                        for k,v in row.items():
                            set_value(response, k, v)
                    dd = device_defaults.get(sot_key)
                    if dd and dd == row.get(csv_key):
                        del row[csv_key]
                        for k,v in row.items():
                            set_value(response, k, v)
    csvfile.close()
    return response

def add_values_from_excel(response, item_config, device_facts, device_defaults, onboarding_config):

    global _global_cache

    table = []

    basedir = "%s/%s" % (onboarding_config.get('git').get('app_configs').get('path'),
                         onboarding_config.get('git').get('app_configs').get('subdir'))
    directory = os.path.join(basedir, './onboarding/required/')

    filename = "%s/%s" % (directory, item_config.get('file'))
    matching_key = item_config.get('matches_on')
    logging.debug(f'reading additional values from {filename}')

    if filename in _global_cache:
        workbook = _global_cache.get(filename)
    else:
        # Load the workbook
        workbook = load_workbook(filename = filename)
        _global_cache[filename] = workbook

    # Select the active worksheet
    worksheet = workbook.active
    
    # loop through table and build list of dict
    rows = worksheet.max_row
    columns = worksheet.max_column + 1 
    for row in range(2, rows + 1):
        line = {}
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)
    
    for row in table:
        # maybe there are multiple items
        for matches_on in item_config.get('matches_on'):
            # sot key => name of key in our sot
            # csv_key => name of column in our csv file
            for sot_key, csv_key in matches_on.items():
                df = device_facts.get(sot_key,'')
                if len(df) > 0 and df.lower() == row.get(csv_key,'').lower():
                    del row[csv_key]
                    for k,v in row.items():
                        # do not add None or empty values
                        if v and len(v) > 0:
                            set_value(response, k, v)
                dd = device_defaults.get(sot_key,'')
                if len(dd) > 0 and dd.lower() == row.get(csv_key,'').lower():
                    del row[csv_key]
                    for k,v in row.items():
                        # do not add None or empty values
                        if v and len(v) > 0:
                            set_value(response, k, v)

def process_matches(response, device_facts, device_defaults, item_config, ciscoconf):
    matches = get_matches(device_facts, 
                          device_defaults, 
                          item_config.get('matches',{}), 
                          ciscoconf)
    if not matches:
        return
    # build dict using key and values configured in 'values'
    for key, value in item_config.get('values').items():
        # logging.debug(f'key {key} value {value}')
        if isinstance(value, str):
            if '__named__' in value:
                group = value.split('__named__')[1]
                response[key] = matches.groups(group)[0]
            else:
                response[key] = value
        elif isinstance(value, dict):
            response[key] = {}
            for k, v in value.items():
                #logging.debug(f'k {k} v {v}')
                if '__named__' in v:
                    groups = v.split('__named__')
                    string = ""
                    for g in groups:
                        if len(g) == 0:
                            continue
                        if g in matches.groupdict():
                            string += matches.group(g)
                        else:
                            string += g
                    response[key][k] = string
                else:
                    response[key][k] = v
                if k == "slug":
                    response[key][k] = slugify(response[key][k])
        else:
            response[key] = value

def get_matches(device_facts, device_defaults, matches, ciscoconf):
    for name, value in matches.items():
        if '__' in name:
            splits = name.split('__')
            source = key = lookup = ""
            if len(splits) == 3:
                # source / key / lookup
                source = splits[0]
                key = splits[1]
                lookup = splits[2]
            elif len(splits) == 2:
                source = splits[0]
                key = splits[1]
            # logging.debug(f'source: {source} key: {key} lookup: {lookup}')

            if source == "facts":
                obj = device_facts.get(key)
            elif source == "defaults":
                obj = device_defaults.get(key)
            elif source == "config":
                # look if value is found in config
                if lookup != "":
                    match = "match__%s" % lookup
                else:
                    match = "match"
                props = {match: value, 'ignore_leading_spaces: True': True}
                if key == "global" and ciscoconf:
                    return ciscoconf.find_in_global(props)
                elif key == "interfaces" and ciscoconf:
                    return ciscoconf.find_in_interfaces(props)
                else:
                    logging.error(f'unknown key; must be global or interfaces')
                    continue
            else:
                logging.error(f'no source found or source {source} invalid')
                continue

            if lookup == '':
                if obj == value:
                    # logging.debug(f'exact match on {key}')
                    return obj
            elif 'ci' == lookup or 'ic' == lookup:
                # logging.debug(f'ci lookup found on {key}')
                if value.lower() in obj.lower():
                    return obj
            elif 're' == lookup or 'rei' == lookup:
                # logging.debug(f'regular expression {value} on {key} found')
                if obj is not None:
                    if 'rei' == lookup:
                        p = re.compile(value, re.IGNORECASE)
                        m = p.search(obj)
                    else:
                        p = re.compile(value)
                        m = p.search(obj)
                    if m:
                        #logging.debug(f'regular expression matches on {obj}')
                        return m
    return False

def read_file(filename, device_defaults):
    with open(filename) as f:
        config = {}
        logging.debug("opening file %s to read required field config" % filename)
        try:
            config = yaml.safe_load(f.read())
            if config is None:
                logging.error("could not parse file %s" % filename)
                return None
        except Exception as exc:
            logging.error("could not read file %s; got exception %s" % (filename, exc))
            return None
        name = config.get('name')
        platform = config.get('platform')

        if not config.get('active'):
            logging.debug(f'file {filename} is not active')
            return None
        if platform is not None:
            if platform != 'all' and platform != device_defaults.get("platform",''):
                logging.debug("skipping custom field %s wrong platform %s" % (name, platform))
                return None

        return config

def set_value(mydict, paths, value):
    # write value to nested dict
    # we split the path by using '__'
    parts = paths.split('__')
    for part in parts[0:-1]:
        # add {} if item does not exists
        # this loop create an empty path
        mydict = mydict.setdefault(part, {})
    # at last write value to dict
    mydict[parts[-1]] = value
