#!/usr/bin/env python

import argparse
import json
import yaml
import pandas as pd
from loguru import logger
from benedict import benedict
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font


def read_xlsx(filename):
    logger.debug(f"Reading file {filename}")
    table = []
    workbook = load_workbook(filename = filename, read_only=True)
    worksheet = workbook.active
    # loop through table and build list of dict
    rows = worksheet.max_row
    columns = worksheet.max_column + 1 
    for row in range(2, rows + 1):
        line = benedict(keyattr_dynamic=True)
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)
    return table

def write_to_xlsx(data, filename):

    # write data to xlsx file
    df = pd.DataFrame(data)

    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Properties', startrow=1, header=False, index=False)

    # Get the xlsxwriter workbook and worksheet objects.
    worksheet = writer.sheets["Properties"]

    # Get the dimensions of the dataframe.
    (max_row, max_col) = df.shape

    # Create a list of column headers, to use in add_table().
    column_settings = [{"header": column} for column in df.columns]

    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 0, max_row, max_col - 1, {"columns": column_settings})

    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1, 12)
    writer.close()

def main():
    parser = argparse.ArgumentParser()

    # set the log level and handler
    parser.add_argument('--loglevel', type=str, required=False, help="used loglevel")
    parser.add_argument('--loghandler', type=str, required=False, help="used log handler")
    parser.add_argument('--input', type=str, required=False, help="filename to read from")
    parser.add_argument('--mapping', type=str, required=False, help="read mapping from file")
    parser.add_argument('--output', type=str, required=False, help="filename to write to")
    parser.add_argument('--parent-of-parent', action='store_true', help='add parent of parent')

    # parse arguments
    args = parser.parse_args()

    # output data
    output_data = []
    keypath_data = benedict()

    # open input file
    data = read_xlsx(args.input)

    # read yaml mapping from file if provided
    if args.mapping:
        with open(args.mapping, 'r') as stream:
            try:
                mapping = yaml.safe_load(stream)['mapping']
            except yaml.YAMLError as exc:
                print(exc)
    else:
        mapping = None

    # get list of keys from dict
    key_list = list(data[0].keys())

    # build our data structure
    for row in data:
        keypath_data['.'.join(row.values())] = "x"

    # get keypaths from our data
    keypaths = keypath_data.keypaths(indexes=False)

    # loop through keypaths and build our data structure
    for kp in keypaths:
        kp_splitted = kp.split('.')
        nn_of_elements = len(kp_splitted)
        if nn_of_elements > 1:
            element = kp_splitted[nn_of_elements -1]
            parent = kp_splitted[nn_of_elements - 2]
        else:
            element = kp_splitted[nn_of_elements -1]
            parent = ""
        
        # check if we have a mapping that we have to use
        if mapping:
            location_type = mapping[key_list[nn_of_elements -1]]
        else:
            location_type = key_list[nn_of_elements -1]

        parents = []
        if args.parent_of_parent and len(kp_splitted) > 2:
            # add all elements up to nn_elemenets -3 to parents
            parents = kp_splitted[:nn_of_elements - 1]
            #print(f'element {element} parents {parents}')

        # add data to output data structure
        row = {'name': element, 
               'parent.name': parent,
               'description': element,
               'location_type.name': location_type,
               'status.name': "Active"}

        if parents:
            for i in range(len(parents) - 1, 0, -1):
                col = ""
                for x in range(i, len(parents)):
                    col += "parent."
                if len(col) > 0:
                    col += "name"
                #print(f'element {element} col {col} parents {parents[i]}')
                row.update({col: parents[i]})               
        output_data.append(row)
    
    if args.output:
        write_to_xlsx(output_data, args.output)
    else:
        print(json.dumps(output_data, indent=4))

if __name__ == "__main__":
    main()