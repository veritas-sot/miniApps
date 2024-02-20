import csv
import yaml
import importlib

from benedict import benedict
from loguru import logger
from openpyxl import load_workbook

# veritas
import veritas.plugin
from veritas.tools import tools


####  bulk update

def bulk_update(sot, filename, updater_config, add_missing_data=False, force=False, dry_run=False):
    # get mapping from config
    key_mapping = updater_config.get('mappings',{}).get('keys',{})
    value_mapping = updater_config.get('mappings',{}).get('valaues',{})

    if 'csv' in filename:
        data = read_csv(filename, updater_config, key_mapping, value_mapping)
    elif 'xlsx' in filename:
        data = read_xlsx(filename, key_mapping, value_mapping)

    # Data without modification does not need to be changed unless the user wants it.
    updates = []
    for row in data:
        if not row['checksum'] or force:
            updates.append(row)

    logger.info(f'updating {len(updates)} items; force={force}')
    if len(updates) > 0:
        contains_interface = True if updates[0].get('interface') else False
        contains_primary_ip = True if updates[0].get('primary_ip4') else False
        contains_primary_interface = True if updates[0].get('primary_ip4',{}).get('interfaces') else False

        logger.debug(f'contains_interface={contains_interface} '
                     f'contains_primary_ip={contains_primary_ip} '
                     f'contains_primary_interface={contains_primary_interface} ')

        # do we have to add missing data like missing IP addresses
        # If you change the primary IP address of a device and have not previously 
        # added the IP to nautobot, nautobot will raises an error.

        if contains_primary_ip and add_missing_data:
            for update in updates:
                add_ip_address(sot, update, dry_run)

        if contains_primary_interface and contains_primary_ip:
            for update in updates:
                update_primary_interface(sot, update, dry_run)

        # are we able to make a bulk update by using the ID?
        if 'id' in updates[0]:
            if contains_interface:
                do_update(sot, updates, updater_config, "dcim/interfaces", dry_run)
            else:
                do_update(sot, updates, updater_config, "dcim/devices", dry_run)
        else:
            # we do not have an ID; use hostname/interface name instead
            if contains_interface:
                for interface in updates:
                    if dry_run:
                        host_name = interface['hostname']
                        interface_name = interface['name']
                        print(f'host {host_name} interface {interface_name} update: {interface}')
                    else:
                        sot.device(interface['hostname']).interface(interface['name']).update(interface)
            else:
                for device in updates:
                    if dry_run:
                        print(f'host {device["hostname"]} update: {device}')
                    else:
                        sot.device(device['hostname']).update(device)

def parse_row(row, key_mapping, value_mapping):
    data = benedict(keyattr_dynamic=True)

    for k, value in row.items():
        if k in key_mapping:
            key = key_mapping.get(k)
        else:
            key = k
        # value_mappings are uses to 'rename' values
        if value_mapping:
            mapping = value_mapping.get(key, {})
            for map_key,map_value in mapping.items():
                if value == map_key:
                    value = map_value

        if key.startswith('cf_'):
            if 'custom_fields' not in data:
                data['custom_fields'] = {}
            data['custom_fields'][key.split('cf_')[1]] = value
        else:
            # we use benedict and the . notation
            data[key] = value
    return data

def read_csv(filename, updater_config, key_mapping={}, value_mapping={}):
    contains_interface = False
    data = []

    delimiter = updater_config['defaults']['import'].get('delimiter',',')
    quotechar = updater_config['defaults']['import'].get('quotechar','|')
    quoting_cf = updater_config['defaults']['import'].get('quoting','minimal')
    newline = updater_config['defaults']['import'].get('newline','')
    if quoting_cf == "none":
        quoting = csv.QUOTE_NONE
    elif quoting_cf == "all":
        quoting = csv.QUOTE_ALL
    elif quoting_cf == "nonnumeric":
        quoting = csv.QUOTE_NONNUMERIC
    else:
        quoting = csv.QUOTE_MINIMAL
    logger.info(f'reading {filename} delimiter={delimiter} quotechar={quotechar} newline={newline} quoting={quoting_cf}')

    # read CSV file
    with open(filename, newline=newline) as csvfile:
        csvreader = csv.DictReader(csvfile, 
                                   delimiter=delimiter, 
                                   quoting=quoting,
                                   quotechar=quotechar)
        for row in csvreader:
            # check if we have interfaces to update or import
            for name in row:
                if 'interface' in name:
                    contains_interface = True

            old_checksum = new_checksum = 0
            if 'checksum' in row:
                old_checksum = row['checksum']
                del row['checksum']
                new_checksum = tools.calculate_md5(list(row.values()))
                logger.debug(f'old_checksum: {old_checksum} new_checksum: {new_checksum}')

            row['checksum'] = old_checksum == new_checksum
            data.append(parse_row(row, key_mapping, value_mapping))

    return contains_interface, data

def read_xlsx(filename, key_mapping={}, value_mapping={}):
    data = []
    table = []

    # Load the workbook
    workbook = load_workbook(filename)
    # Select the active worksheet
    worksheet = workbook.active

    # loop through table and build list of dict
    rows = worksheet.max_row
    # the +1 is important otherwise we miss the the last column (eg. checksum)
    columns = worksheet.max_column + 1 
    for row in range(2, rows + 1):
        line = {}
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)

    for row in table:
        old_checksum = new_checksum = 0
        if any(d == 'checksum' for d in row.keys()):
            old_checksum = row['checksum']
            del row['checksum']
            new_checksum = tools.calculate_md5(list(row.values()))
            logger.debug(f'old_checksum: {old_checksum} new_checksum: {new_checksum}')

        row['checksum'] = old_checksum == new_checksum
        parsed = parse_row(row, key_mapping, value_mapping)
        data.append(parsed)
    return data

def do_update(sot, data, updater_config, endpoint, dry_run):
    if dry_run:
        for d in data:
            print(d)
        return

    nb = sot.rest(url=updater_config['sot']['nautobot'], 
                  token=updater_config['sot']['token'],
                  verify_ssl=updater_config['sot']['ssl_verify'],
                  debug=False)
    nb.session()
    response = nb.patch(url=f"api/{endpoint}/", json=data)
    if response.status_code != 200:
        logger.error(f'could not update data; got error {response.content}')
    else:
        logger.info('data updated')

def add_ip_address(sot, properties, dry_run=False):
    primary_ip = properties.get('primary_ip4.address')
    if not sot.get.address(primary_ip):
        logger.debug(f'IP address {primary_ip} does not exists')
        addr = {'address': primary_ip,
                'status': {'name': 'Active'},
                'namespace': 'Global'
               }
        if dry_run:
            print(f'adding {primary_ip} to IPAM')
            return
        if sot.ipam.add_ip(addr):
            logger.info(f'added {primary_ip} to IPAM')
            return True
        else:
            logger.error(f'could not add IP {primary_ip} to IPAM; this may cause further errors')
            return False

def update_primary_interface(sot, properties, dry_run=False):
    # get id of the device
    # It is best to use the ID. This makes it possible to change all 
    # properties, including the name.
    if 'id' in properties:
        device = sot.get.device(name=properties['id'], by_id=True)
    else:
        device = properties.get('hostname')
    
    if not device:
        logger.error(f'failed to get host {properties.get("hostname")}')
        return False
    # We need the currently configured address
    current_primary_ip = device.primary_ip4

    # get interface
    interface = properties.get('primary_ip4.interfaces.name')

    # remove old assignments
    assignment = sot.ipam.get_assignment(
        interface=interface,
        address=current_primary_ip,
        device=device
        )

    if dry_run:
        print(f'removing current assignment of {interface}')
    else:
        if assignment:
            logger.debug(f'removing current assignment of {interface}')
            try:
                assignment.delete()
            except Exception as exc:
                logger.error(f'failed to delete assignment; got exception {exc}')
                return False
        else:
            logger.debug(f'failed to get assignment of {interface} on {device}')

    primary_ip = properties.get('primary_ip4.address',{})
    interface = properties.get('primary_ip4.interfaces.name')
    assigned = sot.ipam.assign_ipaddress_to_interface(
        device=device,
        interface=interface,
        address=primary_ip
    )
    if dry_run:
        print(f'assigning IP address {primary_ip} to interface {interface}')
    else:
        if assigned:
            logger.debug(f'assigning IP address {primary_ip} to interface {interface}')
            sot.ipam.set_primary(
                device=device,
                address=primary_ip
            )
        else:
            logger.error(f'failed to assign IP address {primary_ip} to interface {interface}')

#### tasks

def run_task(args, sot, job, select, using, where):

    for task in job.get('tasks'):
        device_list = sot.select(select) \
                         .using(using) \
                         .where(where)
        task_name = list(task.keys())[0]
        logger.debug(f'running task {task_name}')
        if task_name in ['add_tag', 'set_tag','delete_tag']:
            tag_management(sot, task_name, task[task_name], device_list, job.get('preprocessing'))
        if 'device_property' in task:
            device_properties(sot, task, device_list, job.get('preprocessing'))
        if 'interface_property' in task:
            interface_properties(sot, task, device_list, job.get('preprocessing'))
        if 'mode' in task and task['mode'] == 'advanced':
            run_advanced_task(sot, task, device_list)

def tag_management(sot, todo, task, device_list, preprocessing):
    scope = task.get('scope')
    configured_tags = task.get('tag', [])
    if isinstance(configured_tags, str):
        tags = [ configured_tags ]
    else:
        tags = configured_tags
    if scope is None or len(tags) == 0:
        logger.error('scope and tags must be configured to set tags')
        return

    for device in device_list:
        name = get_device_name(device, preprocessing)
        if not name:
            # we do not have a name and thus cannot update the device
            continue
        if scope == "dcim.interface":
            for interface in device.get('interfaces', []):
                interface_name = interface.get('name')
                if 'add_tag' == todo:
                    logger.info(f'adding tag {tags} on {name}/{interface_name}')
                    sot.device(name).interface(interface_name).add_tags(tags)
                elif 'set_tag' == todo:
                    logger.info(f'setting tag {tags} on {name}/{interface_name}')
                    sot.device(name).interface(interface_name).set_tags(tags)
                elif 'delete_tag' == todo:
                    logger.info(f'deleting tag {tags} on {name}/{interface_name}')
                    sot.device(name).interface(interface_name).delete_tags(tags)
        elif scope == "dcim.device":
            if 'add_tag' == todo:
                logger.info(f'add tag {tags} on {name}')
                sot.device(name).add_tags(tags)
            elif 'set_tag' == todo:
                logger.info(f'setting tag {tags} on {name}')
                sot.device(name).set_tags(tags)
            elif 'delete_tag' == todo:
                logger.info(f'deleting tag {tags} on {name}')
                sot.device(name).delete_tags(tags)

def device_properties(sot, task, device_list, preprocessing):
    for device in device_list:
        name = get_device_name(device, preprocessing)
        if not name:
            # we do not have a name and thus cannot update the device
            continue

        logger.debug(f'updating {name}')
        # update device
        success = sot.device(name).update(task.get('device_property'))
        if success:
            logger.info(f'updated {name} successfully')
        else:
            logger.info(f'could not update {name}')

def interface_properties(sot, task, device_list, preprocessing):
    for device in device_list:
        name = get_device_name(device, preprocessing)
        for interface in device.get('interfaces', []):
            interface_name = interface.get('name')
            logger.debug(f'updating {name}/{interface_name}')
            success = sot.device(name) \
                         .interface(interface_name) \
                         .update(task.get('interface_property',{}))
            if success:
                logger.info(f'updated {name}/{interface_name} successfully')
            else:
                logger.info(f'could not update {name}/{interface_name}')

def get_device_name(device, preprocessing):
    if preprocessing:
        preprocessed_values = benedict(keyattr_dynamic=True)
        temp = benedict(device, keyattr_dynamic=True)
        for key, key_path in preprocessing.items():
            try:
                preprocessed_values[key] = (temp[key_path])
            except Exception:
                pass
            return preprocessed_values.get('name')
    else:
        return device.get('name')

def run_advanced_task(sot, task, device_list):
    package = task.get('plugin_dir')
    subpackage = task.get('plugin')
    call = task.get('call')
    method_type = task.get('type')

    if not package or not subpackage or not call:
        logger.error('no package, ubpackage or call found')
        return
    
    try:
        importlib.import_module(f'{package}.{subpackage}')
    except Exception as exc:
        logger.critical(f'failed to import plugin {package}.{subpackage}; got exception {exc}')

    plugin = veritas.plugin.Plugin()
    if 'autonomous' == method_type:
        logger.debug(f'calling autonomous method {call}')
        autonomous = plugin.get_kobold_plugin(call)
        if callable(autonomous):
            autonomous(sot=sot, arguments=task.get('arguments'), devices=device_list)
        else:
            logger.error(f'could not call autonomous method {call}')
    elif 'return_value' == method_type:
        logger.debug(f'calling return_value method {call}')
        return_value = plugin.get_kobold_plugin(call)
        if not callable(return_value):
            logger.error(f'could not call return_value method {call}')
            return
        update = task.get('update')

        #
        # loop through device list and update property
        #
        for device_properties in device_list:
            value = return_value(
                sot=sot, 
                arguments=task.get('arguments'), 
                device_properties=device_properties)

            #
            # update entity
            #
            if 'device' in update:
                # we need the ID or the name of the device
                id = device_properties.get('id')
                name = device_properties.get('name', device_properties.get('hostname'))
                if not id and not name:
                    logger.error('failed to update data; no name or id found')
                    return None
                # entity is for debugging purposes
                entity = name if name else id

                update_property = update.get('device')
                logger.debug(f'setting {update_property} to {value} on {entity}')
                if update_property.startswith('cf_'):
                    logger.debug('updating custom_field')
                    success = sot.device(name).update({
                                'custom_fields': 
                                {update_property.replace("cf_",''): value}
                    })
                else:
                    success = sot.device(name).update({update_property: value})
                if success:
                    logger.info(f'successfully updated {update_property} on {entity}')
                else:
                    logger.error(f'failed to updated {update_property} on {entity}')
                

#### main

def do_jobs_from_file(args, sot, updater_config):
    jobs_yaml = read_yaml(args.filename)
    if not jobs_yaml:
        logger.error(f'failed to read YAML config {args.filename}')
        return False
    else:
        jobs = jobs_yaml.get('update')

    for job in jobs:
        if args.job and jobs.get('job') != args.job:
            continue

        # defaults should be
        # devices: name
        # prefixes: prefix, ip_addresses, primary_ip4_for, name
        # addresses: primary_ip4_for, name

        # get select, using, and where
        select = job.get('devices',{}).get('select')
        using =  job.get('devices',{}).get('from')
        where =  args.where if args.where else job.get('devices',{}).get('where')
        logger.debug(f'select={select} where={where} using={using}')

        # we do NOT want to update all devices
        if not where:
            logger.error('there is no where clause specified. We do not want to update ALL devices')
            logger.error('Please use --devices name= if you realy want to update alles devices')
            print('there is no where clause specified. We do not want to update ALL devices')
            print('Please use --devices name= if you realy want to update alles devices')
            return
        
        run_task(args, sot, job, select, using, where)

def read_yaml(filename):
    try:
        with open(filename) as f:
            return yaml.safe_load(f.read())
    except Exception as exc:
        logger.error(f'could not read or parse config; got exception {exc}')
        return None

def update(sot, args, kobold_config):

    if '.csv' in args.filename or '.xlsx' in args.filename:
        bulk_update(
            sot, 
            args.filename, 
            kobold_config, 
            args.add_missing_data, 
            args.force, 
            args.dry_run)
    else:
        do_jobs_from_file(args,sot,kobold_config.get('update',{}))
