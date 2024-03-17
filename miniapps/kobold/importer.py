import json
import re
from loguru import logger
from benedict import benedict
from openpyxl import load_workbook


def read_json(filename):
    logger.debug(f'reading JSON from {filename}')
    try:
        with open(filename, 'r') as f:
            data = json.load(f)
        return benedict(data, keyattr_dynamic=True)
    except Exception as exc:
        logger.error(f'failed to load JSON; got exception {exc}')
        return {}

def read_xlsx(filename):
    table = []
    workbook = load_workbook(filename = filename, read_only=True)
    worksheet = workbook.active
    # loop through table and build list of dict
    rows = worksheet.max_row
    columns = worksheet.max_column + 1 
    for row in range(2, rows + 1):
        line = benedict(keyattr_dynamic=True)
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)
    return table

def read_yaml(filename):
    logger.debug(f'reading yaml {filename}')
    try:
        return benedict(filename, format="yaml")
    except Exception as exc:
        raise Exception (f'could not parse yaml file {filename}; got {exc}')

def import_hldm(sot, device_properties, dry_run=False):

    list_of_interfaces = []

    saved_values = benedict(keyattr_dynamic=True)
    for item in ['interfaces','config_context','primary_ip4','tags', 'hostname']:
        if item in dict(device_properties):
            saved_values[item] = device_properties[item]
            del device_properties[item]

    # remove empty and null values
    device_properties.clean()

    # prepare interfaces
    if 'interfaces' in saved_values:
        for interface in saved_values['interfaces']:
            iface = benedict(interface, keyattr_dynamic=True)
            iface.clean()
            if 'type' in iface:
                iface['type'] = iface['type'].replace('A_','').replace('_','-').lower()
            list_of_interfaces.append(iface)
    else:
        list_of_interfaces.append({})

    # get primary ip
    primary_interface = saved_values['primary_ip4.interfaces[0].name']

    # set tags
    list_of_tags = []
    tags = saved_values.get('tags')
    if tags:
        if isinstance(tags, str):
            # excel sheet
            list_of_tags = tags.split(',')
        elif isinstance(tags, list):
            for tag in tags:
                if isinstance(tag, str):
                    list_of_tags.append(tag)
                elif isinstance(tag, dict):
                    # hldm
                    if 'name' in tag:
                        list_of_tags.append(tag['name'])
    
    device_to_nautobot(sot, device_properties, list_of_interfaces, primary_interface, list_of_tags, dry_run)

def import_ipaddresses(sot, ipaddresses, dry_run=False):
    if dry_run:
        for ip in ipaddresses:
            print(f'importing {ip}')
        return
    success = sot.ipam.add_ip(ipaddresses)
    if success:
        logger.info('successfully imported IP-addresses')
    else:
        logger.error('failed to import IP-addresses')

def import_device_from_xlsx(sot, filename, dry_run=False):

    logger.debug(f'reading workbook {filename}')
    workbook = load_workbook(filename=filename, read_only=True)
    device_sheet = workbook['Device']
    interfaces_sheet = workbook['Interfaces']

    device = benedict(keyattr_dynamic=True)
    device_rows = device_sheet.max_row
    device_header = [None] * device_rows

    interface = benedict(keyattr_dynamic=True)
    list_of_interfaces = []
    interfaces_columns = interfaces_sheet.max_column
    interfaces_rows = interfaces_sheet.max_row
    interfaces_header = [None] * interfaces_columns

    # get device headers
    for row in range(1, device_rows):
        device_header[row-1] = device_sheet.cell(row=row+1, column=1).value

    for row in range(1, device_rows):
        key = device_header[row-1]
        value = device_sheet.cell(row=row+1, column=2).value
        if 'face' == key:
            value = value.lower()
        elif 'vrfs' == key:
            # value looks like vrf_name(vrf_namespace)
            match = re.match("(.*?)\((.*?)\)", value)
            if match:
                vrf_name = match.groups(1)[0]
                vrf_namespace = match.groups(1)[1]
                value = [{'name': vrf_name, 'namespace': vrf_namespace}]
        logger.debug(f'(dev) key={key} value={value}')
        device[key] = value

    logger.configure(extra={"extra": device.get('name')})
    logger.info('importing device')

    # clean empty values
    device.clean()

    # get tags
    if 'tags' in device:
        list_of_tags = device.get('tags').split(',')
        del device['tags']
    else:
        logger.debug(f'device {device["name"]} has no tags')
        list_of_tags = []

    # get interface headers
    for col in range(0, interfaces_columns):
        interfaces_header[col] = interfaces_sheet.cell(row=1, column=col+1).value
    # get primary_interface
    if 'primary_ip4' in device:
        primary_interface = device.get('primary_ip4.interfaces[0].name','primary interface')
        del device['primary_ip4']

    for row in range(2, interfaces_rows + 1):
        interface = benedict(keyattr_dynamic=True)
        for col in range(0, interfaces_columns):
            key = interfaces_header[col]
            value = interfaces_sheet.cell(row=row, column=col+1).value
            if 'type' == key:
                value = value.lower().replace('a_','').replace('_','-')
            elif 'mode' == key and value:
                value = value.lower().replace('_','-')
            elif 'ip_addresses[x].address' == key:
                if value:
                    list_of_ips = value.replace(' ','').split(',')
                    x = 0
                    for ip in list_of_ips:
                        interface[f'ip_addresses[{x}].address'] = ip
                        x += 1
                else:
                    continue
            if value:
                logger.debug(f'(iface) key={key} value={value}')
                interface[key] = value
        interface.clean()
        list_of_interfaces.append(interface)
    
    if dry_run:
        print(f'device: {device.get("name")}')
        print(json.dumps(device, indent=4))
        print(f'{len(list_of_interfaces)} interfaces')
        print(json.dumps(list_of_interfaces, indent=4))
        return

    device_to_nautobot(sot, device, list_of_interfaces, primary_interface, list_of_tags, dry_run)

def device_to_nautobot(sot, device_properties, list_of_interfaces, primary_interface, list_of_tags, dry_run):

    if dry_run:
        name = device_properties.get('name')
        print(f'importing device {name} properties={device_properties} ' \
              f'interfaces={list_of_interfaces} primary={primary_interface}')
    else:
        name = device_properties.get('name')
        logger.debug(f'importing name={device_properties.get("name")} ' \
                     f'primary_interface={primary_interface}')
        new_device = sot.onboarding \
                        .interfaces(list_of_interfaces) \
                        .primary_interface(primary_interface) \
                        .add_prefix(False) \
                        .add_device(device_properties)
        if new_device:
            logger.info(f'imported {name} to nautobot')
        else:
            logger.error(f'failed to import {name} to nautobot')
            return

    # set tags
    if dry_run:
        print(f'adding tags {list_of_tags}')
    else:
        sot.device(new_device.display).set_tags(list_of_tags)

def import_custom_fields(sot, data, dry_run=False):
    set_defaults = []
    for cf in data['custom_fields']:
        logger.debug(f'importing {cf}')
        if 'default' in cf:
            logger.debug('found default value; setting default after creating cf')
            properties = {'label': cf.get('label'), 'default': cf.get('default')}
            set_defaults.append(properties)
            del cf['default']

    logger.debug('import custom fields')
    sot.importer.add(properties=data['custom_fields'], endpoint='custom_fields')
    logger.debug('import custom fields choices')
    sot.importer.add(properties=data['custom_field_choices'], endpoint='custom_field_choices')

    # set default value
    for properties in set_defaults:
        sot.updater.update(endpoint='custom_fields', getter={'label': properties.get('label')}, values=properties)

def import_device_from_yaml(sot, devices, dry_run=False):
    for device in devices['devices']:
        import_hldm(sot, device, dry_run)

def remove_empty_parents(d):
    for k, v in list(d.items()):
        if isinstance(v, dict):
            remove_empty_parents(v)
        if not v:
            del d[k]
    return d

def import_data(sot, args):
    if args.filename and 'json' in args.filename:
        data = read_json(args.filename)
        # check which data we have
        if all(k in data for k in ('name','status','device_type', 'role')):
            import_hldm(sot, data)
    elif args.filename and 'yaml' in args.filename:
        data = read_yaml(args.filename)
        key = list(data.keys())[0]
        logger.debug(f'found key={key}')
        if key in ['location_types','locations', 'roles', 'device_types', 'platforms', 'manufacturers', 
                   'tags', 'webhooks', 'prefixes', 'custom_links']:
            if args.dry_run:
                print(f'importing {data[key]} to {key}')
                return
            success = sot.importer.add(properties=data[key], endpoint=key)
            if success:
                logger.info(f'{key} successfully imported')
            else:
                logger.error(f'failed to import {key}')
        elif key in ['custom_field_choices','custom_fields']:
            import_custom_fields(sot, data)
        elif key == 'devices':
            import_device_from_yaml(sot, data, args.dry_run)
        else:
            logger.error(f'unknown key {key}')
    elif 'xlsx' in args.filename:
        logger.debug(f'reading xlsx file {args.filename}')
        data = read_xlsx(args.filename)

        #
        # IP addresses
        #
        if all(k in data[0] for k in ('address','namespace', 'status')):
            import_ipaddresses(sot, data, args.dry_run)
        
        #
        # import HLDM
        #
        elif all(k in data[0] for k in ('name','status','device_type', 'role')):
            for device in data:
                import_hldm(sot, device, args.dry_run)

        #
        # import single device
        #
        elif all(k in list(data[0].keys()) for k in ('Property','Value')):
             import_device_from_xlsx(sot, args.filename, args.dry_run)

        #
        # import nautobot default values like locations
        #
        elif all(k in data[0] for k in ('name','parent.name','location_type.name','status.name')):
            # this is an xlsx file containing locations
            for item in data:
                # we have to remove empty parents
                # work with a copy of the dict and remove if parent == None
                for key, value in dict(item).items():
                    if key == 'parent':
                        # we have to remove empty parents
                        d = remove_empty_parents(value)
                        if len(d) == 0:
                            del item['parent']
            success = sot.importer.add(properties=data, endpoint='locations')
            if success:
                logger.info('locations successfully imported')
            else:
                logger.error('failed to import locations')
        elif args.endpoint:
            success = sot.importer.add(properties=data, endpoint=args.endpoint)
            if success:
                logger.info(f'{args.endpoint} successfully imported')
            else:
                logger.error(f'failed to import {args.endpoint}')
        else:
            logger.error('could not detect what data it is. Please use --endpoint to import data')
    