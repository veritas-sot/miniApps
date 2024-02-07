import os
import json
import csv
import pandas as pd
import re
from loguru import logger
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from benedict import benedict

# veritas
from veritas.devicemanagement import scrapli as dm
from veritas.tools import tools

def get_profile(profile, username, password, playbook):
    """set username and password in playbook

    Parameters
    ----------
    profile : str
        commandline profile
    username : str
        commandline username
    password : str
        commandline password
    playbook : str
        name of playbook
    """
    BASEDIR = os.path.abspath(os.path.dirname(__file__))

    # check if .env file exists and read it
    if os.path.isfile(os.path.join(BASEDIR, '.env')):
        logger.debug('reading .env file')
        load_dotenv(os.path.join(BASEDIR, '.env'))
    else:
        logger.debug('no .env file found; trying to read local crypto parameter')
        crypt_parameter = tools.get_miniapp_config('onboarding', BASEDIR, "salt.yaml")
        os.environ['ENCRYPTIONKEY'] = crypt_parameter.get('crypto', {}).get('encryptionkey')
        os.environ['SALT'] = crypt_parameter.get('crypto', {}).get('salt')
        os.environ['ITERATIONS'] = str(crypt_parameter.get('crypto', {}).get('iterations'))

    # load profiles
    profile_config = tools.get_miniapp_config('kobold', BASEDIR, 'profiles.yaml')
    # get username and password either from profile
    username, password = tools.get_username_and_password(
            profile_config,
            profile,
            username,
            password)

    playbook.username = username
    playbook.password = password

def get_value(values, keys):
    if isinstance(values, list):
        my_list = []
        for value in values:
            my_list.append(get_value(value.get(keys[0]), keys[1:]))
        return my_list
    elif isinstance(values, str):
        return values
    if len(keys) == 1:
        if values is None:
            return ''
        else:
            return values.get(keys[0])
    return get_value(values.get(keys[0]), keys[1:])

def get_number_of_rows(raw_data):
    rows = -1
    for key, values in raw_data.items():
        if isinstance(values, list):
            if rows == -1:
                rows = len(values)
            else:
                # we have multiple lists and the number of rows
                # differ. In this case we are not able to export the data
                if rows != len(values):
                    logger.error('the number of rows differ for the results')
                    return None
    return rows

def get_device_config_and_facts(sot, playbook, device_properties):
    device_facts = {}

    device_ip = device_properties.get('primary_ip4',{}).get('address')
    # check if device_ip is cidr notation
    device_ip=device_ip.split('/')[0]
    platform = device_properties.get('platform',{}).get('name')
    # sometimes devices have no manufacturer
    manfctrr = device_properties.get('platform',{}).get('manufacturer',{})
    if not manfctrr or manfctrr == 'None':
        logger.info('this device has no manufacturer; using cisco')
        manufacturer = 'cisco'
    else:
        manufacturer = manfctrr.get('name','cisco')

    # Get the path to the directory this file is in
    BASEDIR = os.path.abspath(os.path.dirname(__file__))
    # Connect the path with the '.env' file name
    load_dotenv(os.path.join(BASEDIR, '.env'))

    logger.info(f'ssh to {playbook.username}@{device_ip}:{playbook.tcp_port} on platform {platform}')
    conn = dm.Devicemanagement(ip=device_ip,
                               platform=platform,
                               manufacturer=manufacturer.lower(),
                               username=playbook.username,
                               password=playbook.password,
                               port=playbook.tcp_port)

    # retrieve facts like fqdn, model and serialnumber
    logger.debug(f'now gathering facts from {device_ip}')
    device_facts = conn.get_facts()
    if device_facts is None:
        logger.error('got no facts; skipping device')
        if conn:
            conn.close()
        return None, None
    device_facts['args.device'] = device_ip

    # retrieve device config
    logger.info("getting running-config")
    try:
        device_config = conn.get_config("running-config")
    except Exception as exc:
        logger.error("could not receive device config from %s; got exception %s" % (device_ip, exc))
        return None, None
    if device_config is None:
        logger.error(f'could not retrieve device config from {device_ip}')
        conn.close()
        return None, None
    conn.close()

    return device_config, device_facts

def get_device_data_to_export(sot, task, devices):

    """
    prepare data to export by scanning through ALL devices and build table
    """

    # our list of dicts that contains the data we export
    data_to_export = []
    # some defaults
    calculate_checksum = False
    header_written = False

    # columns is the list of device/interface properties the user wants to export
    columns = task.get('columns').replace(' ','').split(',')

    for device in devices:
        device_property = {}
        for column in columns:
            data = get_value(device, column.split('.'))
            logger.debug(f'column={column} data={data}')
            device_property[column] = data
        
        number_of_rows = get_number_of_rows(device_property)
        if not number_of_rows:
            logger.error('number of rows are different for some columns')
            continue
        logger.debug(f'number_of_rows={number_of_rows}')

        # first add header if user wants it
        if 'header' in task and not header_written:
            header_written = True
            row = []
            for column in columns:
                row.append(column)
            data_to_export.append(row)

        # loop through the number of rows
        for n in range(number_of_rows):
            # initialize empty row
            row = [None] * len(columns)
            # and set column to 0
            number_of_column = 0
            for column in columns:
                # check if we have to calculate an MD5 sum
                if 'checksum' in column:
                    calculate_checksum = True
                if isinstance(device_property[column], str):
                    row[number_of_column] = device_property[column]
                else:
                    if device_property[column] is None:
                        row[number_of_column] = 'null'
                    else:
                        dta =  device_property[column][n]
                        if isinstance(dta, list):
                            if len(dta) > 0:
                                row[number_of_column] = dta[0]
                            else:
                                row[number_of_column] = ""
                        else:
                            row[number_of_column] = dta
                number_of_column += 1
            # calculate checksum if necessary
            if calculate_checksum:
                row[len(columns)-1] = tools.calculate_md5(row)

            # now put row to the list of exported values
            data_to_export.append(row)
    return data_to_export

def export_config_and_facts(sot, playbook, task, device_properties):

    hostname = device_properties.get('hostname')
    logger.debug(f'getting config and facts from {hostname}')
    device_config, device_facts = get_device_config_and_facts(sot, playbook, device_properties)
    content = task.get('content')
    BASEDIR = os.path.abspath(os.path.dirname(__file__))

    if 'config' in content:
        filename = "%s/%s/%s.conf" % (
            BASEDIR, 
            task.get('directory','./configs'), 
            hostname)
        subdir = os.path.dirname(filename)
        if not os.path.exists(subdir):
                logger.info(f'creating missing directory {subdir}')
                os.makedirs(subdir)

        logger.info(f'writing config to {filename}')
        try:
            with open(filename, 'w') as f:
                f.write(device_config)
        except Exception as exc:
            logger.error(f'could not write config; got exception {exc}')

    if 'facts' in content:
        hostname = device_properties.get('hostname')
        filename = "%s/%s/%s.facts" % (
            BASEDIR, 
            task.get('directory','./configs'), 
            hostname)
        subdir = os.path.dirname(filename)
        if not os.path.exists(subdir):
                logger.info(f'creating missing directory {subdir}')
                os.makedirs(subdir)

        logger.info(f'writing facts to {filename}')
        try:
            with open(filename, 'w') as f:
                f.write(json.dumps(device_facts,indent=4))
        except Exception as exc:
            logger.error(f'could not write facts; got exception {exc}')

def export_hldm(sot, playbook, task, devices):
    filename_pattern = task.get('filename', '__name__')
    subdir_pattern = task.get('directory', '')

    for device in devices:
        hostname = device.get('name')
        logger.debug(f'exporting HLDM of {hostname}')
        hldm = sot.get.hldm(device=hostname, get_id=False)[0]
        subdir = playbook.pattern_to_filename(subdir_pattern, hldm)
        filename = playbook.pattern_to_filename(filename_pattern, hldm)

        logger.debug(f'writing HLDM to subdir {subdir} filename {filename}')
        if not os.path.exists(subdir):
                logger.info(f'creating missing directory {subdir}')
                os.makedirs(subdir)
        with open(f'{subdir}/{filename}', "w") as f:
            f.write(json.dumps(hldm, indent=4))

#
# device properties
#

def export_device_properties_as_csv(task, data_to_export):
    logger.info(f'exporting {len(data_to_export)} entries as CSV')
    delimiter = task.get('delimiter',',')
    quotechar = task.get('quotechar','|')
    quoting_cf = task.get('quoting', 'minimal')
    filename = task.get('filename','export.csv')

    if quoting_cf == "none":
        quoting = csv.QUOTE_NONE
    elif quoting_cf == "all":
        quoting = csv.QUOTE_ALL
    elif quoting_cf == "nonnumeric":
        quoting = csv.QUOTE_NONNUMERIC
    else:
        quoting = csv.QUOTE_MINIMAL

    # create directory if it does not exsists
    directory = os.path.dirname(filename)
    if not os.path.exists(directory):
        os.makedirs(directory)

    logger.info(f'exporting data to {filename}')

    # now write our csv file
    with open(filename, 'w', newline='') as csvfile:
        export_writer = csv.writer(csvfile, delimiter=delimiter, quotechar=quotechar, quoting=quoting)
        for line in data_to_export:
            export_writer.writerow(line)

def export_device_properties_as_xlsxl(task, data_to_export):
    filename = task.get('filename','export.xlsx')
    logger.bind(extra='exp properties').info(f'exporting data as EXCEL to {filename}')

    # create directory if it does not exsists
    directory = os.path.dirname(filename)
    if not os.path.exists(directory):
        os.makedirs(directory)

    df = pd.DataFrame(data_to_export)
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Export', startrow=0, header=False, index=False)
    writer.close()

def export_device_properties(sot, playbook, task, devices):
    data_to_export = get_device_data_to_export(sot, task, devices)
    if len(data_to_export) == 0:
        logger.bind(extra='export').info('got no data to export')
        return
    if task.get('format') == 'csv':
        return export_device_properties_as_csv(task, data_to_export)
    if task.get('format') == 'excel' or task.get('format') == 'xlsx':
        return export_device_properties_as_xlsxl(task, data_to_export)

#
# export devcice to xlsx
#

def export_device_to_xlsx(sot, playbook, task, devices):

    # get the columns we have to export
    columns = task.get('columns')
    colors = task.get('colors')

    for dev in devices:
        # we need a benedict to get all values of the device
        device = benedict(dev, keyattr_dynamic=True)
        #print(json.dumps(device, indent=4))
        name = device.get('name')
        logger.configure(extra={"extra": name})

        # create workbook and sheets (Device - renamed - , Interfaces)
        workbook = Workbook()
        device_sheet = workbook.active
        device_sheet.title = "Device"
        interfaces_sheet = workbook.create_sheet("Interfaces")
        interfaces_sheet.title = "Interfaces"
        interfaces_sheet.sheet_properties.tabColor = "1072BA"

        # we need two lists
        list_of_tags = []
        list_of_vrfs = []

        # add header and style cells
        row = 2
        ft = Font(color=colors.get('header_font', 'FFFFFF'))
        header_color = colors.get('header','004c81ba')
        a1 = device_sheet["A1"]
        b1 = device_sheet["B1"]
        a1.value = "Property"
        b1.value = "Value"
        a1.font = ft
        b1.font = ft
        a1.fill = PatternFill(
                    start_color=header_color, end_color=header_color, fill_type = "solid")
        b1.fill = PatternFill(
                    start_color=header_color, end_color=header_color, fill_type = "solid")

        # add device data to 'Device' sheets
        for property in columns['device']:
            value = device.get(property)
            if value:
                if 'tags' == property:
                    for tag in value:
                        list_of_tags.append(tag['name'])
                    value = ",".join(list_of_tags)
                elif 'vrfs' == property:
                    for vrf in value:
                        vrf_name = vrf.get('name','')
                        vrf_namespace = vrf.get('namespace',{}).get('name','')
                        list_of_vrfs.append(f'{vrf_name}({vrf_namespace})')
                    value = ",".join(list_of_vrfs)
                elif isinstance(value, list) and len(value) == 0:
                    continue
                logger.debug(f'key={property} value={value}')
                device_sheet.cell(column=1, row=row).value = property
                device_sheet.cell(column=2, row=row).value = value
                color = colors.get(property, colors.get('default','00FFFFFF'))
                device_sheet.cell(column=1, row=row).fill = PatternFill(
                        start_color=color, end_color=color, fill_type = "solid")
                device_sheet.cell(column=2, row=row).fill = PatternFill(
                        start_color=color, end_color=color, fill_type = "solid")
                row += 1

        # get list of interfaces
        interfaces = device.get('interfaces',[])

        # add header to interface sheet
        # this list is dynamic beacause some interfaces may
        # have multiple IP addresses and for each address we have to add one column
        interface_headers = []

        # add header to interfaces sheet
        
        for property in columns['interfaces']:
            interface_headers.append(property)
        interfaces_sheet.append(interface_headers)

        # add data to interface_values
        for iface in interfaces:
            interface = benedict(iface, keyattr_dynamic=True)
            values = []
            for property in columns['interfaces']:
                # some properties are lists
                # ip_addresses[x].address or tagged_vlans[x].vid
                # we get the values and concat them to a ','.join string
                match = re.match("(.*?)\[x\]\.(.*)", property)
                if match:
                    logger.debug(f'found [x] identifier in {property}')
                    list_item = match.groups(1)[0]
                    sub_item = match.groups(1)[1]
                    items = interface[list_item]
                    list_of_values = []
                    for item in items:
                        list_of_values.append(str(item[sub_item]))
                    value = ','.join(list_of_values)
                else:
                    try:
                        value = interface[property]
                    except Exception:
                        value = ''
                logger.debug(f'key={property} value={value}')
                values.append(value)
            interfaces_sheet.append(values)
  
        # and format interface sheet as table
        interfaces_col = chr(64+len(interface_headers))
        interfaces_row = len(interfaces) + 1
        logger.debug(f'interface table: A1:{interfaces_col}{interfaces_row}')
        interface_table = Table(displayName="Interfaces", ref=f'A1:{interfaces_col}{interfaces_row}')
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False)
        interface_table.tableStyleInfo = style
        interfaces_sheet.add_table(interface_table)

        set_column_width(device_sheet, 1.0)
        set_column_width(interfaces_sheet)

        # create directory if it does not exsists and save it
        filename = task.get('filename').replace('__name__', name)
        directory = os.path.dirname(filename)
        if not os.path.exists(directory):
            os.makedirs(directory)

        workbook.save(filename=filename)

def set_column_width(sheet, factor=1.1):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2) * factor
        sheet.column_dimensions[column_letter].width = adjusted_width

#
# main
#

def run_task(args, sot, playbook, tasks, devices):
    for task in tasks:
        content = task.get('content')
        if 'config' in content or 'facts' in content:

            # set username and password
            get_profile(args.profile, args.username,args.password, playbook)

            for device in  devices:
                export_config_and_facts(sot, playbook, task, device)
        elif 'hldm' in content:
            export_hldm(sot, playbook, task, devices)
        elif 'properties' in content:
            # export columns of all devices
            export_device_properties(sot, playbook, task, devices)
        elif 'device_to_xlsx' in content:
            export_device_to_xlsx(sot, playbook, task, devices)

def export(sot, playbook, args):
    job = playbook.jobs.get(args.job)
    if not job:
        logger.error(f'unknown job {args.job}')
        return False

    name = job.get('job')
    description = job.get('description','no description')
    logger.info(f'starting job {name} / {description}')

    if 'sql' in job.get('devices',{}):
        sql = job.get('devices').get('sql')
        select = sql.get('select')
        using = sql.get('from', sql.get('using'))
        where = sql.get('where')
        logger.debug(f'getting device_list select={select} using={using} where={where}')
        device_list = sot.select(select) \
                         .using(using) \
                         .where(where)
        logger.info(f'got {len(device_list)} devices')
    tasks = job.get('tasks')
    if tasks is None:
        logger.error('no task configured!!!')
        return False

    for task in tasks:
        if 'export' in task:
            run_task(args, sot, playbook, task['export'], device_list)

