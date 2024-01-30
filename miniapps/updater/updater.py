#!/usr/bin/env python

import argparse
import os
import json
import csv
import yaml
import urllib3
import re
import jinja2
from benedict import benedict
from loguru import logger
from pathlib import Path
from openpyxl import load_workbook

import veritas.logging
from veritas.sot import sot as sot
from veritas.tools import tools


def build_dict(value, my_dict, keys):
    if isinstance(keys, list):
        my_dict[keys[0]] = build_dict(value, my_dict, keys[1])
    else:
        return {keys: value}
    return my_dict

def parse_row(row, key_mapping, value_mapping):
    data = {}

    for k, value in row.items():
        if k in key_mapping:
            key = key_mapping.get(k)
        else:
            key = k
        # value_mappings are uses to 'rename' values
        if value_mapping:
            mapping = value_mapping.get(key, {})
            for map_key,map_value in mapping.items():
                if value == map_key:
                    value = map_value

        if key.startswith('cf_'):
            if 'custom_fields' not in data:
                data['custom_fields'] = {}
            data['custom_fields'][key.split('cf_')[1]] = value
        elif '__' in key:
            # check if key has subkeys
            path = key.split('__')
            d = build_dict(value, {}, path)
            if path[0] == 'interfaces':
                data.update(d[path[0]])
            else:
                if path[0] not in data:
                    data[path[0]] = {}
                data[path[0]].update(d[path[0]])
        else:
            data[key] = value
    return data

def read_csv(filename, updater_config, key_mapping={}, value_mapping={}):
    contains_interface = False
    data = []

    delimiter = updater_config['defaults']['import'].get('delimiter',',')
    quotechar = updater_config['defaults']['import'].get('quotechar','|')
    quoting_cf = updater_config['defaults']['import'].get('quoting','minimal')
    newline = updater_config['defaults']['import'].get('newline','')
    if quoting_cf == "none":
        quoting = csv.QUOTE_NONE
    elif quoting_cf == "all":
        quoting = csv.QUOTE_ALL
    elif quoting_cf == "nonnumeric":
        quoting = csv.QUOTE_NONNUMERIC
    else:
        quoting = csv.QUOTE_MINIMAL
    logger.info(f'reading {filename} delimiter={delimiter} quotechar={quotechar} newline={newline} quoting={quoting_cf}')

    # read CSV file
    with open(filename, newline=newline) as csvfile:
        csvreader = csv.DictReader(csvfile, 
                                   delimiter=delimiter, 
                                   quoting=quoting,
                                   quotechar=quotechar)
        for row in csvreader:
            # check if we have interfaces to update or import
            for name in row:
                if 'interface' in name:
                    contains_interface = True

            old_checksum = new_checksum = 0
            if 'checksum' in row:
                old_checksum = row['checksum']
                del row['checksum']
                new_checksum = tools.calculate_md5(list(row.values()))
                logger.debug(f'old_checksum: {old_checksum} new_checksum: {new_checksum}')

            row['checksum'] = old_checksum == new_checksum
            data.append(parse_row(row, key_mapping, value_mapping))

    return contains_interface, data

def read_xlsx(filename, key_mapping={}, value_mapping={}):
    contains_interface = False
    data = []
    table = []

    # Load the workbook
    workbook = load_workbook(filename)
    # Select the active worksheet
    worksheet = workbook.active
    
    # loop through table and build list of dict
    rows = worksheet.max_row
    # the +1 is important otherwise we miss the the last column (eg. checksum)
    columns = worksheet.max_column +1 
    for row in range(2, rows + 1):
        line = {}
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)

    for row in table:
        old_checksum = new_checksum = 0
        contains_interface = any(d.startswith('interface') for d in row.keys())
        if any(d == 'checksum' for d in row.keys()):
            old_checksum = row['checksum']
            del row['checksum']
            new_checksum = tools.calculate_md5(list(row.values()))
            logger.debug(f'old_checksum: {old_checksum} new_checksum: {new_checksum}')

        row['checksum'] = old_checksum == new_checksum
        data.append(parse_row(row, key_mapping, value_mapping))
    logger.debug(f'contains_interface={contains_interface}')
    return contains_interface, data

def do_update(sot, data, updater_config, endpoint, dry_run):
    if dry_run:
        for d in data:
            print(d)
        return

    nb = sot.rest(url=updater_config['sot']['nautobot'], 
                  token=updater_config['sot']['token'],
                  verify_ssl=updater_config['sot']['ssl_verify'])
    nb.session()
    response = nb.patch(url=f"api/{endpoint}/", json=data)
    if response.status_code != 200:
        logger.error(f'could not update data; got error {response.content}')
    else:
        logger.info(f'data updated')

def bulk_update(sot, filename, updater_config, dry_run=False):
    # get mapping from config
    key_mapping = updater_config.get('mappings',{}).get('keys',{})
    value_mapping = updater_config.get('mappings',{}).get('valaues',{})

    if 'csv' in filename:
        contains_interface, data = read_csv(filename, updater_config, key_mapping, value_mapping)
    elif 'xlsx' in filename:
        contains_interface, data = read_xlsx(filename, key_mapping, value_mapping)

    # Data without modification does not need to be changed unless the user wants it.
    updates = []
    for row in data:
        if not row['checksum'] or args.force:
            updates.append(row)

    logger.info(f'{len(updates)} to be updated force={args.force} interfaces: {contains_interface}')
    if len(updates) > 0:
        # are we able to make a bulk update by using the ID?
        if 'id' in updates[0]:
            if contains_interface:
                do_update(sot, updates, updater_config, "dcim/interfaces", dry_run)
            else:
                do_update(sot, updates, updater_config, "dcim/devices", dry_run)
        else:
            # we do not have an ID; use hostname/interface name instead
            if contains_interface:
                for interface in updates:
                    if dry_run:
                        host_name = interface['hostname']
                        interface_name = interface['name']
                        print(f'host {host_name} interface {interface_name} update: {interface}')
                    else:
                        sot.device(interface['hostname']).interface(interface['name']).update(interface)
            else:
                for device in updates:
                    if dry_run:
                        print(f'host {device["hostname"]} update: {device}')
                    else:
                        sot.device(device['hostname']).update(device)

def camel(s):
  s = re.sub(r"(_|-)+", " ", s).title().replace(" ", "")
  return ''.join([s[0].lower(), s[1:]])

def get_value_from_template(device, template):
    # read template
    with open(template) as f:
        template = f.read()
    j2 = jinja2.Environment(loader=jinja2.BaseLoader, trim_blocks=False).from_string(template)
    try:
        return j2.render({'values': device})
    except Exception as exc:
        logger.error("could not render template; got exception: %s" % exc)

    return ""

def update_from_file(sot, filename, where, template, updater_config, using='nb.devices', dry_run=False):
    """read config from file and update items depending on this config"""

    # the left part is the item the right part a modifier like upper or lower
    modifier = re.compile("__(.*?)@(.*?)__")
    zfill = re.compile("__(.*?)@zfill\((\d+)\)__")

    # init vars
    named_groups = {}
    destinations = {}

    with open(filename) as f:
        try:
            config = yaml.safe_load(f.read())
        except Exception as exc:
            logger.error(f'could not read or parse config {exc}')
            return None

    # compile named groups
    cfg = config.get('update', {}).get('source', {}).get('named_groups','')
    for item in cfg:
        logger.debug(f'named: {item} pattern: {cfg[item]}')
        named_groups[item] = re.compile(cfg[item])

    cfg = config.get('update', {}).get('destination', {})
    for item in cfg:
        logger.debug(f'destination: {item} new value: {cfg[item]}')
        destinations[item] = cfg[item]

    # get items to update
    select = set()
    select.add('id')
    if using == 'nb.devices':
        for ng in named_groups:
            select.add(ng.split('.')[0])
    elif using == "nb.ipaddresses":
        select.add('address')
        select.add('interface_assignments')

    for key in named_groups.keys():
        select.add(key)
    itemlist = sot.select(list(select)) \
                  .using(using) \
                  .where(where)

    if len(itemlist) == 0 and dry_run:
        print('nothing to do')
        return

    logger.info(f'got {len(itemlist)} item from our sot')

    # loop through items and check if it must be updated
    for row in itemlist:

        # we use benedict
        # the advantage is that we can easily get values from it
        # entity is the row we got from our SOT
        # this can be a device or an ip address
        entity = benedict(row, keyattr_dynamic=True)
        updates = benedict(keyattr_dynamic=True)

        if using == 'nb.devices':
            extra = entity['hostname']
            hostname_id = entity['id']
        elif using == "nb.ipaddresses":
            extra = entity['address']
            address_id = entity['id']
        else:
            logger.error(f'unknown or unsupported type {using}')
            return None

        # loop through named groups and check if pattern matches
        # matched_values contains all items that were found in entity
        # and for which the pattern matched
        matched_values = {}
        for ng_key, pattern in named_groups.items():
            # we have to check if the key can be found in our entity
            try:
                item = entity[ng_key]
            except KeyError as e:
                logger.error(f'key {ng_key} not found in entity')
                continue
            logger.bind(extra=extra).debug(f'key: {ng_key} pattern: {pattern} item: {item}')
            match = pattern.match(item)
            if match:
                for group, group_val in match.groupdict().items():
                    matched_values[group] = group_val

        if len(matched_values) == 0:
            logger.bind(extra=extra).debug(f'entity without matching group')

        # now matched_values is complete
        # we loop through the destinations and set the new value
        for parameter, orig_value in destinations.items():
            new_value = orig_value
            logger.debug(f'parameter={parameter} new_value from config: {new_value}')
            for group, group_val in matched_values.items():
                # check if we have to fill up a named group
                # this is a special case because we have an argument
                match = zfill.match(new_value)
                if match:
                    item = match.group(1)
                    fill = match.group(2)
                    new_value = new_value.replace(f'__{group}@zfill({fill})__', group_val.zfill(int(fill)))

                match = modifier.match(new_value)
                # at first check if we have to use a emplate
                if template:
                    logger.debug(f'using template to get new_value')
                    print(row)
                    new_value = get_value_from_template(row, template)
                # now check if we have some modifiers (upper, lower)
                elif match:
                    item = match.group(1)
                    mod = match.group(2)
                    logger.debug(f'item={item} modifier={mod}')
                    if 'upper' == mod:
                        new_value = new_value.replace(f'__{group}@upper__', group_val.upper())
                    elif 'lower' == mod:
                        new_value = new_value.replace(f'__{group}@lower__', group_val.lower())
                    elif 'title' == mod:
                        new_value = new_value.replace(f'__{group}@title__', group_val.title())
                    elif 'capwords' == mod:
                        new_value = new_value.replace(f'__{group}@capwords__', group_val.capwords())
                    elif 'camel':
                        new_value = new_value.replace(f'__{group}@cammel__', camel(group_val))
                    else:
                        logger.error(f'unknown mod value {mod}')
                # otherwise replace named groups in new_value
                else:
                    new_value = new_value.replace(f'__{group}__', group_val)

                logger.bind(extra=extra).debug(f'parameter: {parameter} group {group} '\
                    f'group_val: {group_val} final new_value: {new_value}')
                if parameter.startswith('cf_'):
                    parameter = parameter.replace('cf_','')
                    if 'custom_fields' in updates:
                        updates['custom_fields'].update({parameter: new_value})
                    else:
                        updates['custom_fields'] = {parameter: new_value}
                else:
                    # because we use benedict we can easily set the new value
                    # the syntax of the key is like key.subkey.subsubkey
                    # eg location.location_type.name
                    updates[parameter] = new_value

        # now we are able to update the item in nautobot
        if len(updates) > 0:
            if using == 'nb.devices':
                if dry_run:
                    print(f'update {extra}; new values: {updates}')
                else:
                    try:
                        nb_obj = sot.get.device(name=hostname_id, by_id=True)
                        response = nb_obj.update(data=updates)
                        logger.bind(extra=extra).info(f'item updated; data={updates}; response={response}')
                    except Exception as exc:
                        logger.bind(extra=extra).error(f'could not update item {exc}')
            elif using == 'nb.ipaddresses':
                if dry_run:
                    print(f'update {extra}; new values: {updates}')
                else:
                    try:
                        nb_obj = sot.get.address(address=address_id, by_id=True)
                        response = nb_obj.update(data=updates)
                        logger.bind(extra=extra).info(f'item updated; data={updates}; response={response}')
                    except Exception as exc:
                        logger.bind(extra=extra).error(f'could not update item')


if __name__ == "__main__":
    # to disable warning if TLS warning is written to console
    urllib3.disable_warnings()

    parser = argparse.ArgumentParser()

    # the user can enter a different config file
    parser.add_argument('--config', type=str, required=False, help="updater config file")
    # what devices to update
    # this parameter is only used in if --update was set
    parser.add_argument('--devices', type=str, required=False, help="query to get list of devices")
    # use address to update IP addresses
    parser.add_argument('--addresses', type=str, required=False, help="query to get list of IP addresses")
    # set the log level and handler
    parser.add_argument('--loglevel', type=str, required=False, help="used loglevel")
    parser.add_argument('--loghandler', type=str, required=False, help="used log handler")
    # uuid is written to the database logger
    parser.add_argument('--uuid', type=str, required=False, help="database logger uuid")

    parser.add_argument('--bulk-update', type=str, required=False, help="use file to update data")
    parser.add_argument('--update', type=str, required=False, help="use yaml config to update data")
    parser.add_argument('--template', type=str, default="", required=False, help="template to use to update value")
    # force is only used with bulk-update
    parser.add_argument('--force', action='store_true', help='force update even if checksum is equal')
    parser.add_argument('--dry-run', action='store_true', help='print updates only')

    # parse arguments
    args = parser.parse_args()

    # Get the path to the directory this file is in
    BASEDIR = os.path.abspath(os.path.dirname(__file__))
    
    # read config
    updater_config = tools.get_miniapp_config('updater', BASEDIR, args.config)
    if not updater_config:
        print('unable to read config')
        sys.exit()

    # create logger environment
    veritas.logging.create_logger_environment(
        config=updater_config, 
        cfg_loglevel=args.loglevel,
        cfg_loghandler=args.loghandler,
        app='updater',
        uuid=args.uuid)

    # we need the SOT object to talk to the SOT
    sot = sot.Sot(token=updater_config['sot']['token'],
                  ssl_verify=updater_config['sot'].get('ssl_verify', False),
                  url=updater_config['sot']['nautobot'])

    if args.bulk_update:
        bulk_update(sot, args.bulk_update, updater_config, args.dry_run)
    if args.update and args.devices:
        update_from_file(sot, args.update, args.devices, args.template, updater_config, 'nb.devices', args.dry_run)
    if args.update and args.addresses:
        update_from_file(sot, args.update, args.addresses, None, updater_config, 'nb.ipaddresses', args.dry_run)
